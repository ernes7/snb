import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# Colors
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(bold=True, size=11, color="1F4E79")
ERNESTO_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
JUNIOR_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

ernesto_teams = ["Granma", "Sancti Spíritus", "Pinar del Río", "Las Tunas"]
junior_teams = ["Ciego de Ávila", "Villa Clara", "Santiago de Cuba", "Industriales"]
all_teams = ernesto_teams + junior_teams

draft_order = ["Granma", "Las Tunas", "Pinar del Río", "Ciego de Ávila",
               "Sancti Spíritus", "Santiago de Cuba", "Villa Clara", "Industriales"]
draft_ratings = {
    "Granma": 28, "Las Tunas": 26, "Pinar del Río": 25, "Ciego de Ávila": 23,
    "Sancti Spíritus": 18, "Santiago de Cuba": 15, "Villa Clara": 13, "Industriales": 3
}
post_draft_ratings = {
    "Granma": 22, "Las Tunas": 18, "Pinar del Río": 17, "Ciego de Ávila": 11,
    "Sancti Spíritus": 5, "Santiago de Cuba": 2, "Villa Clara": 3, "Industriales": 4
}

# Draft results: team -> [(player_name, position, original_team), ...]
draft_picks = {
    "Granma":          [("R. Videaux", "RF", ""), ("Y. Paumier", "2B", ""), ("M. Gonzales", "SP", "")],
    "Las Tunas":       [("O. Arias", "C", ""), ("G. Duvergel", "CF", ""), ("M. Lahera", "SP", "")],
    "Pinar del Río":   [("O. Del Rosario", "LF", ""), ("Y. Cerce", "2B", ""), ("Miranda Jr", "SP", "")],
    "Ciego de Ávila":  [("M. Enriguez", "3B", ""), ("A. Sanchez", "LF", ""), ("Y. Pedroso", "SP", "")],
    "Sancti Spíritus": [("D. Moreira", "SS", ""), ("Ibanez", "RF", ""), ("D. Hinojosa", "SP", "")],
    "Santiago de Cuba": [("Orta", "LF", ""), ("J. Abreu", "1B", ""), ("C. Yanes", "SP", "")],
    "Villa Clara":     [("Vido", "LF", ""), ("Garcia", "1B", ""), ("J. Martinez", "SP", "")],
    "Industriales":    [("Pacheco", "SS", ""), ("Jose Angel Garcia", "SP", ""), ("Y. Gonzales", "SP", "")],
}

def get_player(team):
    return "Ernesto" if team in ernesto_teams else "Junior"

def get_team_fill(team):
    return ERNESTO_FILL if team in ernesto_teams else JUNIOR_FILL

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

def style_data_cell(ws, row, col, center=True):
    cell = ws.cell(row=row, column=col)
    cell.border = THIN_BORDER
    if center:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    return cell

# ============================================================
# SHEET 1: DRAFT
# ============================================================
ws_draft = wb.active
ws_draft.title = "Draft"

ws_draft.merge_cells('A1:F1')
ws_draft['A1'] = "DRAFT - Torneo MVP Cuba 2011"
ws_draft['A1'].font = TITLE_FONT

ws_draft.merge_cells('A2:F2')
ws_draft['A2'] = "Reglas: Cada equipo elige 3 jugadores. Máximo 2 lanzadores o 2 bateadores (no 3 del mismo grupo)."
ws_draft['A2'].font = Font(italic=True, size=10, color="666666")

ws_draft.merge_cells('A3:F3')
ws_draft['A3'] = "Pool: Jugadores de los 8 equipos no seleccionados"
ws_draft['A3'].font = Font(italic=True, size=10, color="666666")

# Draft order reference
row = 5
ws_draft.merge_cells(f'A{row}:F{row}')
ws_draft[f'A{row}'] = "Orden del Draft (peor primero)"
ws_draft[f'A{row}'].font = SUBTITLE_FONT

row = 6
headers = ["Orden", "Equipo", "Jugador", "OVR Pre", "OVR Post", "Cambio"]
for col, h in enumerate(headers, 1):
    ws_draft.cell(row=row, column=col, value=h)
style_header_row(ws_draft, row, len(headers))

for i, team in enumerate(draft_order):
    r = row + 1 + i
    style_data_cell(ws_draft, r, 1).value = i + 1
    cell = style_data_cell(ws_draft, r, 2, center=False)
    cell.value = team
    cell.fill = get_team_fill(team)
    style_data_cell(ws_draft, r, 3).value = get_player(team)
    style_data_cell(ws_draft, r, 4).value = draft_ratings[team]
    style_data_cell(ws_draft, r, 5).value = post_draft_ratings[team]
    change = draft_ratings[team] - post_draft_ratings[team]
    change_cell = style_data_cell(ws_draft, r, 6)
    change_cell.value = f"+{change}" if change > 0 else str(change)
    change_cell.font = Font(bold=True, color="008000" if change > 0 else "FF0000")

# Draft picks table
row = 16
ws_draft.merge_cells(f'A{row}:F{row}')
ws_draft[f'A{row}'] = "Selecciones del Draft"
ws_draft[f'A{row}'].font = SUBTITLE_FONT

row = 17
pick_headers = ["Pick #", "Equipo", "Jugador Elegido", "Posición", "Equipo Original", "Notas"]
for col, h in enumerate(pick_headers, 1):
    ws_draft.cell(row=row, column=col, value=h)
style_header_row(ws_draft, row, len(pick_headers))

# 24 picks (8 teams x 3 rounds) - filled with results
pick_num = 0
for rnd in range(3):
    for team in draft_order:
        pick_num += 1
        r = row + pick_num
        style_data_cell(ws_draft, r, 1).value = pick_num
        cell = style_data_cell(ws_draft, r, 2, center=False)
        cell.value = team
        cell.fill = get_team_fill(team)
        name, pos, orig = draft_picks[team][rnd]
        style_data_cell(ws_draft, r, 3).value = name
        style_data_cell(ws_draft, r, 4).value = pos
        style_data_cell(ws_draft, r, 5).value = orig
        style_data_cell(ws_draft, r, 6)

# Column widths
ws_draft.column_dimensions['A'].width = 8
ws_draft.column_dimensions['B'].width = 20
ws_draft.column_dimensions['C'].width = 22
ws_draft.column_dimensions['D'].width = 12
ws_draft.column_dimensions['E'].width = 20
ws_draft.column_dimensions['F'].width = 20

# ============================================================
# SHEET 2: STANDINGS
# ============================================================
ws_stand = wb.create_sheet("Clasificación")

ws_stand.merge_cells('A1:I1')
ws_stand['A1'] = "CLASIFICACIÓN - Torneo MVP Cuba 2011"
ws_stand['A1'].font = TITLE_FONT

row = 3
stand_headers = ["Equipo", "Jugador", "G", "P", "PCT", "DIF GB", "CA", "CL", "DIFF"]
for col, h in enumerate(stand_headers, 1):
    ws_stand.cell(row=row, column=col, value=h)
style_header_row(ws_stand, row, len(stand_headers))

for i, team in enumerate(all_teams):
    r = row + 1 + i
    cell = style_data_cell(ws_stand, r, 1, center=False)
    cell.value = team
    cell.fill = get_team_fill(team)
    style_data_cell(ws_stand, r, 2).value = get_player(team)

    # G (Wins) - COUNTIF on Regular Season where team is home and home_runs > away_runs, or away and away_runs > home_runs
    team_ref = f'"{team}"'
    # Using formulas that reference the Regular Season sheet
    # Win as home: home team matches AND home runs > away runs
    # Win as away: away team matches AND away runs > home runs
    w_formula = (
        f'=COUNTIFS(\'Temporada Regular\'!$C$3:$C$98,{team_ref},\'Temporada Regular\'!$E$3:$E$98,">"&\'Temporada Regular\'!$F$3:$F$98)'
        f'+COUNTIFS(\'Temporada Regular\'!$D$3:$D$98,{team_ref},\'Temporada Regular\'!$F$3:$F$98,">"&\'Temporada Regular\'!$E$3:$E$98)'
    )
    style_data_cell(ws_stand, r, 3).value = w_formula

    l_formula = (
        f'=COUNTIFS(\'Temporada Regular\'!$C$3:$C$98,{team_ref},\'Temporada Regular\'!$F$3:$F$98,">"&\'Temporada Regular\'!$E$3:$E$98)'
        f'+COUNTIFS(\'Temporada Regular\'!$D$3:$D$98,{team_ref},\'Temporada Regular\'!$E$3:$E$98,">"&\'Temporada Regular\'!$F$3:$F$98)'
    )
    style_data_cell(ws_stand, r, 4).value = l_formula

    # PCT
    pct_cell = style_data_cell(ws_stand, r, 5)
    pct_cell.value = f'=IF((C{r}+D{r})=0,"",C{r}/(C{r}+D{r}))'
    pct_cell.number_format = '0.000'

    # GB - relative to first row (will need manual sort)
    style_data_cell(ws_stand, r, 6).value = f'=IF(ROW()=4,"",((C$4-D$4)-(C{r}-D{r}))/2)'

    # Runs scored (CA)
    ca_formula = (
        f'=SUMPRODUCT((\'Temporada Regular\'!$C$3:$C$98={team_ref})*\'Temporada Regular\'!$E$3:$E$98)'
        f'+SUMPRODUCT((\'Temporada Regular\'!$D$3:$D$98={team_ref})*\'Temporada Regular\'!$F$3:$F$98)'
    )
    style_data_cell(ws_stand, r, 7).value = ca_formula

    # Runs allowed (CL)
    cl_formula = (
        f'=SUMPRODUCT((\'Temporada Regular\'!$C$3:$C$98={team_ref})*\'Temporada Regular\'!$F$3:$F$98)'
        f'+SUMPRODUCT((\'Temporada Regular\'!$D$3:$D$98={team_ref})*\'Temporada Regular\'!$E$3:$E$98)'
    )
    style_data_cell(ws_stand, r, 8).value = cl_formula

    # DIFF
    style_data_cell(ws_stand, r, 9).value = f'=G{r}-H{r}'

ws_stand.column_dimensions['A'].width = 20
ws_stand.column_dimensions['B'].width = 12
for c in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
    ws_stand.column_dimensions[c].width = 10

# ============================================================
# SHEET 3: REGULAR SEASON
# ============================================================
ws_reg = wb.create_sheet("Temporada Regular")

ws_reg.merge_cells('A1:M1')
ws_reg['A1'] = "TEMPORADA REGULAR - 96 Juegos"
ws_reg['A1'].font = TITLE_FONT

row = 2
reg_headers = ["#", "Fecha", "Local", "Visitante", "C Local", "C Visit",
               "H Local", "H Visit", "E Local", "E Visit",
               "Pitcher G", "Pitcher P", "Notas"]
for col, h in enumerate(reg_headers, 1):
    ws_reg.cell(row=row, column=col, value=h)
style_header_row(ws_reg, row, len(reg_headers))

# Generate schedule: each Ernesto team vs each Junior team, 6 games (3H, 3A)
schedule = []
for e_team in ernesto_teams:
    for j_team in junior_teams:
        # 3 games where e_team is home
        for _ in range(3):
            schedule.append((e_team, j_team))
        # 3 games where j_team is home
        for _ in range(3):
            schedule.append((j_team, e_team))

random.seed(2011)  # Reproducible shuffle
random.shuffle(schedule)

for i, (home, away) in enumerate(schedule):
    r = row + 1 + i
    style_data_cell(ws_reg, r, 1).value = i + 1
    style_data_cell(ws_reg, r, 2)  # date blank
    cell_home = style_data_cell(ws_reg, r, 3, center=False)
    cell_home.value = home
    cell_home.fill = get_team_fill(home)
    cell_away = style_data_cell(ws_reg, r, 4, center=False)
    cell_away.value = away
    cell_away.fill = get_team_fill(away)
    for c in range(5, 14):
        style_data_cell(ws_reg, r, c)

ws_reg.column_dimensions['A'].width = 5
ws_reg.column_dimensions['B'].width = 12
ws_reg.column_dimensions['C'].width = 20
ws_reg.column_dimensions['D'].width = 20
for c in ['E', 'F', 'G', 'H', 'I', 'J']:
    ws_reg.column_dimensions[c].width = 9
ws_reg.column_dimensions['K'].width = 16
ws_reg.column_dimensions['L'].width = 16
ws_reg.column_dimensions['M'].width = 20

# ============================================================
# SHEET 4: SEMIFINALS
# ============================================================
ws_semi = wb.create_sheet("Semifinales")

ws_semi.merge_cells('A1:M1')
ws_semi['A1'] = "SEMIFINALES - Mejor de 5"
ws_semi['A1'].font = TITLE_FONT

ws_semi.merge_cells('A2:M2')
ws_semi['A2'] = "Si 2 equipos del mismo jugador se enfrentan, ese jugador elige al ganador."
ws_semi['A2'].font = Font(italic=True, size=10, color="666666")

# Series A
row = 4
ws_semi[f'A{row}'] = "Serie A"
ws_semi[f'A{row}'].font = SUBTITLE_FONT

row = 5
ws_semi[f'A{row}'] = "Equipo 1:"
ws_semi[f'B{row}'] = ""  # fill in later
ws_semi[f'D{row}'] = "Equipo 2:"
ws_semi[f'E{row}'] = ""

row = 6
ws_semi[f'A{row}'] = "Serie:"
ws_semi[f'B{row}'] = ""  # e.g. "3-1"

row = 8
semi_headers = ["Juego", "Fecha", "Local", "Visitante", "C Local", "C Visit",
                "H Local", "H Visit", "E Local", "E Visit",
                "Pitcher G", "Pitcher P", "Notas"]
for col, h in enumerate(semi_headers, 1):
    ws_semi.cell(row=row, column=col, value=h)
style_header_row(ws_semi, row, len(semi_headers))

for g in range(5):
    r = row + 1 + g
    style_data_cell(ws_semi, r, 1).value = g + 1
    for c in range(2, 14):
        style_data_cell(ws_semi, r, c)

# Series B
row = 16
ws_semi[f'A{row}'] = "Serie B"
ws_semi[f'A{row}'].font = SUBTITLE_FONT

row = 17
ws_semi[f'A{row}'] = "Equipo 1:"
ws_semi[f'B{row}'] = ""
ws_semi[f'D{row}'] = "Equipo 2:"
ws_semi[f'E{row}'] = ""

row = 18
ws_semi[f'A{row}'] = "Serie:"
ws_semi[f'B{row}'] = ""

row = 20
for col, h in enumerate(semi_headers, 1):
    ws_semi.cell(row=row, column=col, value=h)
style_header_row(ws_semi, row, len(semi_headers))

for g in range(5):
    r = row + 1 + g
    style_data_cell(ws_semi, r, 1).value = g + 1
    for c in range(2, 14):
        style_data_cell(ws_semi, r, c)

ws_semi.column_dimensions['A'].width = 10
ws_semi.column_dimensions['B'].width = 16
ws_semi.column_dimensions['C'].width = 20
ws_semi.column_dimensions['D'].width = 20
for c in ['E', 'F', 'G', 'H', 'I', 'J']:
    ws_semi.column_dimensions[c].width = 9
ws_semi.column_dimensions['K'].width = 16
ws_semi.column_dimensions['L'].width = 16
ws_semi.column_dimensions['M'].width = 20

# ============================================================
# SHEET 5: FINALS
# ============================================================
ws_final = wb.create_sheet("Final")

ws_final.merge_cells('A1:M1')
ws_final['A1'] = "FINAL - Mejor de 5"
ws_final['A1'].font = TITLE_FONT

row = 3
ws_final[f'A{row}'] = "Equipo 1:"
ws_final[f'B{row}'] = ""
ws_final[f'D{row}'] = "Equipo 2:"
ws_final[f'E{row}'] = ""

row = 4
ws_final[f'A{row}'] = "Serie:"
ws_final[f'B{row}'] = ""

row = 6
final_headers = ["Juego", "Fecha", "Local", "Visitante", "C Local", "C Visit",
                 "H Local", "H Visit", "E Local", "E Visit",
                 "Pitcher G", "Pitcher P", "Notas"]
for col, h in enumerate(final_headers, 1):
    ws_final.cell(row=row, column=col, value=h)
style_header_row(ws_final, row, len(final_headers))

for g in range(5):
    r = row + 1 + g
    style_data_cell(ws_final, r, 1).value = g + 1
    for c in range(2, 14):
        style_data_cell(ws_final, r, c)

# Champion
row = 13
ws_final.merge_cells(f'A{row}:M{row}')
champ_cell = ws_final[f'A{row}']
champ_cell.value = "CAMPEÓN:"
champ_cell.font = Font(bold=True, size=16, color="C00000")
champ_cell.alignment = Alignment(horizontal='center')

ws_final.column_dimensions['A'].width = 10
ws_final.column_dimensions['B'].width = 16
ws_final.column_dimensions['C'].width = 20
ws_final.column_dimensions['D'].width = 20
for c in ['E', 'F', 'G', 'H', 'I', 'J']:
    ws_final.column_dimensions[c].width = 9
ws_final.column_dimensions['K'].width = 16
ws_final.column_dimensions['L'].width = 16
ws_final.column_dimensions['M'].width = 20

# ============================================================
# FULL ROSTER DATA
# ============================================================

# Rankings: team -> (overall, pitching, batting, fielding, speed)
team_rankings = {
    "Sancti Spíritus": (5, 13, 5, 10, 7),
    "Granma":          (22, 18, 17, 26, 25),
    "Pinar del Río":   (17, 17, 22, 11, 2),
    "Las Tunas":       (18, 23, 14, 17, 6),
    "Santiago de Cuba": (2, 8, 3, 5, 3),
    "Villa Clara":     (3, 5, 10, 2, 5),
    "Industriales":    (4, 2, 15, 3, 4),
    "Ciego de Ávila":  (11, 4, 20, 9, 8),
}

# Rosters: team -> { lineup: [(pos, name, bats, drafted)], bench: [...], rotation: [...], bullpen: [(role, name, throws)] }
rosters = {
    "Sancti Spíritus": {
        "lineup": [
            ("LF", "F. Cepeda", "L", False),
            ("1B", "Y. Mendoza", "L", False),
            ("RF", "Y. Ibanez", "L", True),
            ("SS", "D. Moreira", "R", True),
            ("3B", "Y. Gourriel", "R", False),
            ("2B", "L. Gourriel Jr", "R", False),
            ("C",  "E. Sanchez", "R", False),
            ("CF", "Y. Gourriel", "R", False),
            ("DH", "Y. Bello", "R", False),
        ],
        "bench": [
            ("2B", "O. Acebey", "R"), ("SS", "Y. Baguet", "A"), ("RF", "L. Monteagudo", "R"),
            ("2B", "J. Alfonso", "R"), ("LF", "A. Gomez", "R"),
        ],
        "rotation": [
            ("I. Jimenez", "R", False), ("D. Hinojosa", "R", True), ("A. Pena", "R", False),
            ("N. Hernandez", "R", False), ("R. Licor", "L", False),
        ],
        "bullpen": [
            ("PRL", "Y. Sosa", "R"), ("PRL", "Y. Panama", "R"), ("PRL", "L. Santana Jr.", "R"),
            ("PRM", "O. Luis Jr", "L"), ("PE", "D. Quintero", "R"),
        ],
    },
    "Granma": {
        "lineup": [
            ("RF", "R. Videaux", "L", True),
            ("LF", "A. Despaigne", "R", False),
            ("3B", "Y. Paumier", "R", True),
            ("CF", "Y. Cespedes", "R", False),
            ("1B", "Y. Samon", "R", False),
            ("DH", "R. Tamayo", "R", False),
            ("SS", "M. Fonseca", "R", False),
            ("2B", "C. Benitez", "R", False),
            ("C",  "L. Ferrales", "R", False),
        ],
        "bench": [
            ("2B", "Y. Milian", "R"), ("LF", "U. Guerra", "R"), ("C", "H. Sanchez", "R"),
            ("2B", "A. Moreno", "R"), ("C", "C. Barrabi Jr.", "R"),
        ],
        "rotation": [
            ("M. Gonzalez", "R", True), ("C. Licea", "R", False), ("M. Vega", "R", False),
            ("A. Soto", "R", False), ("N. Pina", "R", False),
        ],
        "bullpen": [
            ("PRL", "A. Diez", "R"), ("PRL", "J. Pena", "R"), ("PRL", "L. Ramirez", "R"),
            ("PRM", "E. Blanco", "R"), ("PE", "A. Tamayo", "R"),
        ],
    },
    "Pinar del Río": {
        "lineup": [
            ("2B", "Y. Cerce", "R", True),
            ("3B", "D. Duarte", "R", False),
            ("DH", "O. Del Rosario", "R", True),
            ("SS", "D. Castillo", "R", False),
            ("1B", "N. Concepcion", "R", False),
            ("C",  "Y. Peraza", "R", False),
            ("CF", "R. Leon", "R", False),
            ("LF", "M. Rivera", "R", False),
            ("RF", "L. Blanco", "L", False),
        ],
        "bench": [
            ("C", "L. Quintana", "R"), ("3B", "W. Saavedra", "R"), ("SS", "O. Madera", "R"),
            ("C", "O. Pena", "R"), ("SS", "L. Valdes", "R"), ("RF", "L. Gomez", "R"),
        ],
        "rotation": [
            ("G. Miranda Jr.", "R", True), ("V. Banos", "R", False), ("Y. Torres", "R", False),
            ("E. Casanova", "R", False), ("I. Castro", "L", False),
        ],
        "bullpen": [
            ("PRL", "O. Licourt", "R"), ("PRL", "A. Martinez", "L"), ("PRL", "M. Martinez", "R"),
            ("PRM", "J. Martinez", "L"),
        ],
    },
    "Las Tunas": {
        "lineup": [
            ("DH", "G. Duvergel", "L", True),
            ("1B", "J. Pedroso", "R", False),
            ("LF", "Y. Scull", "R", False),
            ("CF", "A. Quiala", "R", False),
            ("2B", "D. Castro", "R", False),
            ("3B", "Y. Alarcon", "R", False),
            ("C",  "O. Arias", "R", True),
            ("RF", "J. Johnson", "L", False),
            ("SS", "A. Guerrero", "R", False),
        ],
        "bench": [
            ("C", "Y. Alarcon", "R"), ("C", "A. Leyva", "R"), ("3B", "M. Brito", "R"),
            ("SS", "D. Rodriguez", "R"), ("CF", "Y. Rondon", "R"),
        ],
        "rotation": [
            ("M. Lahera", "R", True), ("U. Bermudez", "R", False), ("Y. Cruz", "R", False),
            ("D. Mejias", "R", False), ("J. Rojas", "R", False),
        ],
        "bullpen": [
            ("PRL", "Y. Navas", "R"), ("PRL", "E. Sanchez", "R"),
            ("PRM", "A. Guerra", "R"), ("PRM", "Y. Rodriguez", "L"), ("PRM", "J. Moreno", "R"),
        ],
    },
    "Santiago de Cuba": {
        "lineup": [
            ("2B", "H. Olivera", "R", False),
            ("CF", "R. Hurtado", "R", False),
            ("RF", "A. Bell", "R", False),
            ("DH", "J. Abreu", "R", True),
            ("C",  "R. Merino", "R", False),
            ("1B", "P. Poll", "R", False),
            ("SS", "L. Nava", "R", False),
            ("LF", "R. Orta", "R", True),
            ("3B", "M. Castellanos", "L", False),
        ],
        "bench": [
            ("C", "A. Durruthy", "R"), ("LF", "L. Vinent", "R"), ("3B", "Y. Hurtado", "R"),
            ("LF", "L. La O", "R"), ("2B", "D. Aguilera", "R"), ("C", "R. Garcia", "R"),
        ],
        "rotation": [
            ("N. Vera", "R", False), ("O. Romero", "R", False), ("C. Yanes", "R", True),
            ("A. Bicet", "R", False), ("Y. Sanchez", "R", False),
        ],
        "bullpen": [
            ("PRL", "R. Pena", "R"), ("PRL", "E. Valentin", "R"), ("PRL", "D. Betancourt", "R"),
            ("PRM", "O. Tamayo", "R"),
        ],
    },
    "Villa Clara": {
        "lineup": [
            ("RF", "R. Lunar", "R", False),
            ("CF", "A. Zamora", "L", False),
            ("DH", "Y. Garcia", "R", True),
            ("1B", "A. Borrero", "L", False),
            ("C",  "A. Pestano", "R", False),
            ("LF", "Y. Vido", "R", True),
            ("3B", "Y. Canto", "R", False),
            ("2B", "Y. Diaz", "R", False),
            ("SS", "E. Paret", "R", False),
        ],
        "bench": [
            ("RF", "J. Delgado", "R"), ("SS", "A. Diaz", "R"), ("LF", "Y. Flores", "L"),
            ("CF", "A. Garcia", "R"), ("C", "Y. La Rosa", "R"), ("2B", "Y. Perez", "R"),
        ],
        "rotation": [
            ("J. Martinez", "R", True), ("L. Borroto", "R", False), ("F. Alvarez", "R", False),
            ("Y. Lopez", "R", False), ("Y. Perez", "R", False),
        ],
        "bullpen": [
            ("PRL", "R. Carrillo", "L"), ("PRL", "M. Silverio", "R"),
            ("PRM", "Y. Hdez Rojas", "L"), ("PE", "Y. Junquera", "R"),
        ],
    },
    "Industriales": {
        "lineup": [
            ("RF", "S. Hernandez", "L", False),
            ("LF", "I. Chirino", "L", False),
            ("SS", "R. Reyes", "R", False),
            ("1B", "A. Malleta", "R", False),
            ("DH", "Y. Urgelles", "L", False),
            ("C",  "L. Correa", "R", False),
            ("2B", "Y. Pacheco", "R", True),
            ("3B", "J. Torriente", "R", False),
            ("CF", "C. Tabares", "R", False),
        ],
        "bench": [
            ("RF", "S. Perez", "R"), ("3B", "Y. Amador", "R"), ("C", "F. Morejon", "R"),
            ("2B", "R. Olivares", "R"), ("RF", "Y. Tomas", "R"),
        ],
        "rotation": [
            ("Y. Gonzalez", "L", True), ("Y. Marti", "R", False), ("J. Garcia", "R", True),
            ("O. Despaigne", "R", False), ("I. Rendon", "L", False),
        ],
        "bullpen": [
            ("PRL", "A. Rivero", "R"), ("PRL", "F. Montieth", "R"), ("PRL", "A. Sanchez", "L"),
            ("PRM", "G. Concepcion", "L"), ("PE", "A. Romero", "R"),
        ],
    },
    "Ciego de Ávila": {
        "lineup": [
            ("RF", "A. Sanchez", "L", True),
            ("1B", "Y. Charles", "R", False),
            ("3B", "M. Enriquez", "R", True),
            ("CF", "Y. Fiss", "R", False),
            ("DH", "I. Martinez", "L", False),
            ("SS", "Y. Borroto", "R", False),
            ("2B", "M. Vega", "A", False),
            ("LF", "A. Civil", "R", False),
            ("C",  "L. Diaz", "R", False),
        ],
        "bench": [
            ("1B", "Y. Cuesta", "R"), ("2B", "Y. Ramirez", "R"), ("C", "O. Vazquez", "R"),
            ("LF", "A. Jimenez", "R"), ("3B", "R. Gonzalez", "R"),
        ],
        "rotation": [
            ("V. Garcia", "R", False), ("Y. Pedroso", "R", True), ("M. Folch", "L", False),
            ("Y. Guevara", "R", False), ("A. Mora", "R", False),
        ],
        "bullpen": [
            ("PRL", "E. Veliz", "R"), ("PRL", "P. Echemendia", "R"), ("PRL", "D. Duquesne", "R"),
            ("PRM", "I. Gonzalez", "L"), ("PRM", "O. Carrero", "R"),
        ],
    },
}

# ============================================================
# SHEET 6: DATABASE (for dropdowns in other sheets)
# ============================================================
ws_db = wb.create_sheet("Base de Datos")

DRAFTED_FONT = Font(bold=True, color="C00000", size=9)
NORMAL_DB_FONT = Font(size=9)
SECTION_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

ws_db.merge_cells('A1:C1')
ws_db['A1'] = "BASE DE DATOS - Torneo MVP Cuba 2011"
ws_db['A1'].font = TITLE_FONT

# --- Reference lists (columns A-C) ---
row = 3
ws_db.cell(row=row, column=1, value="Equipos")
ws_db.cell(row=row, column=2, value="Jugadores")
ws_db.cell(row=row, column=3, value="Posiciones")
style_header_row(ws_db, row, 3)

for i, team in enumerate(all_teams):
    r = row + 1 + i
    cell = style_data_cell(ws_db, r, 1, center=False)
    cell.value = team
    cell.fill = get_team_fill(team)

style_data_cell(ws_db, 4, 2).value = "Ernesto"
style_data_cell(ws_db, 5, 2).value = "Junior"

positions = ["SP", "RP", "CP", "CE", "C", "1B", "2B", "3B", "SS", "LF", "CF", "RF", "DH",
             "PA", "PRL", "PRM", "PE"]
for i, pos in enumerate(positions):
    style_data_cell(ws_db, row + 1 + i, 3).value = pos

# --- Rankings table (columns A-F, below reference lists) ---
rank_start = row + 1 + max(len(all_teams), len(positions)) + 2
ws_db.merge_cells(f'A{rank_start}:F{rank_start}')
ws_db[f'A{rank_start}'] = "RANKINGS POR EQUIPO"
ws_db[f'A{rank_start}'].font = SUBTITLE_FONT

rh = rank_start + 1
rank_headers = ["Equipo", "Jugador", "OVR", "Pitcheo", "Bateo", "Fildeo", "Velocidad"]
for col, h in enumerate(rank_headers, 1):
    ws_db.cell(row=rh, column=col, value=h)
style_header_row(ws_db, rh, len(rank_headers))

for i, team in enumerate(all_teams):
    r = rh + 1 + i
    cell = style_data_cell(ws_db, r, 1, center=False)
    cell.value = team
    cell.fill = get_team_fill(team)
    style_data_cell(ws_db, r, 2).value = get_player(team)
    ovr, pit, bat, fld, spd = team_rankings[team]
    style_data_cell(ws_db, r, 3).value = ovr
    style_data_cell(ws_db, r, 4).value = pit
    style_data_cell(ws_db, r, 5).value = bat
    style_data_cell(ws_db, r, 6).value = fld
    style_data_cell(ws_db, r, 7).value = spd

ws_db.column_dimensions['A'].width = 20
ws_db.column_dimensions['B'].width = 12
ws_db.column_dimensions['C'].width = 12

# ============================================================
# ONE SHEET PER TEAM with full roster
# ============================================================
for team in all_teams:
    # Sheet name max 31 chars
    short_names = {
        "Sancti Spíritus": "SSP Roster",
        "Granma": "GRA Roster",
        "Pinar del Río": "PRI Roster",
        "Las Tunas": "LTU Roster",
        "Santiago de Cuba": "SCU Roster",
        "Villa Clara": "VCL Roster",
        "Industriales": "IND Roster",
        "Ciego de Ávila": "CAV Roster",
    }
    ws_t = wb.create_sheet(short_names[team])
    roster = rosters[team]
    player_name = get_player(team)
    team_fill = get_team_fill(team)
    ovr, pit, bat, fld, spd = team_rankings[team]

    # Title
    ws_t.merge_cells('A1:F1')
    ws_t['A1'] = f"{team} ({player_name})"
    ws_t['A1'].font = TITLE_FONT

    # Rankings row
    ws_t['A2'] = f"OVR: {ovr}"
    ws_t['B2'] = f"Pitcheo: {pit}"
    ws_t['C2'] = f"Bateo: {bat}"
    ws_t['D2'] = f"Fildeo: {fld}"
    ws_t['E2'] = f"Velocidad: {spd}"
    for c in range(1, 6):
        ws_t.cell(row=2, column=c).font = Font(bold=True, size=10)

    # --- LINEUP ---
    row = 4
    ws_t[f'A{row}'] = "ALINEACIÓN (Orden al Bate)"
    ws_t[f'A{row}'].font = SUBTITLE_FONT
    ws_t[f'A{row}'].fill = SECTION_FILL
    for c in range(2, 7):
        ws_t.cell(row=row, column=c).fill = SECTION_FILL

    row = 5
    lu_headers = ["#", "Pos", "Nombre", "Batea", "Draft", "Notas"]
    for col, h in enumerate(lu_headers, 1):
        ws_t.cell(row=row, column=col, value=h)
    style_header_row(ws_t, row, len(lu_headers))

    for i, (pos, name, bats, drafted) in enumerate(roster["lineup"]):
        r = row + 1 + i
        style_data_cell(ws_t, r, 1).value = i + 1
        style_data_cell(ws_t, r, 2).value = pos
        cell_name = style_data_cell(ws_t, r, 3, center=False)
        cell_name.value = name
        if drafted:
            cell_name.font = DRAFTED_FONT
        style_data_cell(ws_t, r, 4).value = bats
        style_data_cell(ws_t, r, 5).value = "✓" if drafted else ""
        style_data_cell(ws_t, r, 6)

    # --- BENCH ---
    row = 5 + len(roster["lineup"]) + 2
    ws_t[f'A{row}'] = "BANCA"
    ws_t[f'A{row}'].font = SUBTITLE_FONT
    ws_t[f'A{row}'].fill = SECTION_FILL
    for c in range(2, 7):
        ws_t.cell(row=row, column=c).fill = SECTION_FILL

    row += 1
    bench_headers = ["#", "Pos", "Nombre", "Batea", "", ""]
    for col, h in enumerate(bench_headers, 1):
        ws_t.cell(row=row, column=col, value=h)
    style_header_row(ws_t, row, len(bench_headers))

    for i, (pos, name, bats) in enumerate(roster["bench"]):
        r = row + 1 + i
        style_data_cell(ws_t, r, 1).value = i + 1
        style_data_cell(ws_t, r, 2).value = pos
        style_data_cell(ws_t, r, 3, center=False).value = name
        style_data_cell(ws_t, r, 4).value = bats
        style_data_cell(ws_t, r, 5)
        style_data_cell(ws_t, r, 6)

    # --- ROTATION ---
    row = row + len(roster["bench"]) + 2
    ws_t[f'A{row}'] = "ROTACIÓN DE PITCHEO"
    ws_t[f'A{row}'].font = SUBTITLE_FONT
    ws_t[f'A{row}'].fill = SECTION_FILL
    for c in range(2, 7):
        ws_t.cell(row=row, column=c).fill = SECTION_FILL

    row += 1
    rot_headers = ["#", "Rol", "Nombre", "Tira", "Draft", ""]
    for col, h in enumerate(rot_headers, 1):
        ws_t.cell(row=row, column=col, value=h)
    style_header_row(ws_t, row, len(rot_headers))

    for i, (name, throws, drafted) in enumerate(roster["rotation"]):
        r = row + 1 + i
        style_data_cell(ws_t, r, 1).value = i + 1
        style_data_cell(ws_t, r, 2).value = "PA"
        cell_name = style_data_cell(ws_t, r, 3, center=False)
        cell_name.value = name
        if drafted:
            cell_name.font = DRAFTED_FONT
        style_data_cell(ws_t, r, 4).value = throws
        style_data_cell(ws_t, r, 5).value = "✓" if drafted else ""
        style_data_cell(ws_t, r, 6)

    # --- BULLPEN ---
    row = row + len(roster["rotation"]) + 2
    ws_t[f'A{row}'] = "BULLPEN"
    ws_t[f'A{row}'].font = SUBTITLE_FONT
    ws_t[f'A{row}'].fill = SECTION_FILL
    for c in range(2, 7):
        ws_t.cell(row=row, column=c).fill = SECTION_FILL

    row += 1
    bp_headers = ["#", "Rol", "Nombre", "Tira", "", ""]
    for col, h in enumerate(bp_headers, 1):
        ws_t.cell(row=row, column=col, value=h)
    style_header_row(ws_t, row, len(bp_headers))

    for i, (role, name, throws) in enumerate(roster["bullpen"]):
        r = row + 1 + i
        style_data_cell(ws_t, r, 1).value = i + 1
        style_data_cell(ws_t, r, 2).value = role
        style_data_cell(ws_t, r, 3, center=False).value = name
        style_data_cell(ws_t, r, 4).value = throws
        style_data_cell(ws_t, r, 5)
        style_data_cell(ws_t, r, 6)

    # Note about closer
    row = row + len(roster["bullpen"]) + 2
    ws_t[f'A{row}'] = "* El cerrador (CE) no aparece en screenshots pero existe en el equipo."
    ws_t[f'A{row}'].font = Font(italic=True, size=9, color="666666")

    # Column widths
    ws_t.column_dimensions['A'].width = 5
    ws_t.column_dimensions['B'].width = 8
    ws_t.column_dimensions['C'].width = 22
    ws_t.column_dimensions['D'].width = 8
    ws_t.column_dimensions['E'].width = 8
    ws_t.column_dimensions['F'].width = 15

# ============================================================
# ADD DATA VALIDATION (dropdowns) TO OTHER SHEETS
# ============================================================

# Team dropdown for Regular Season (Local & Visitante columns)
team_range = f"'Base de Datos'!$A$4:$A$11"
team_dv = DataValidation(type="list", formula1=team_range, allow_blank=True)
team_dv.prompt = "Seleccione equipo"
team_dv.promptTitle = "Equipo"

# Team dropdowns for Semifinals
team_dv_semi = DataValidation(type="list", formula1=team_range, allow_blank=True)
ws_semi.add_data_validation(team_dv_semi)
team_dv_semi.add('B5')   # Serie A team 1
team_dv_semi.add('E5')   # Serie A team 2
team_dv_semi.add('B17')  # Serie B team 1
team_dv_semi.add('E17')  # Serie B team 2
for g in range(5):
    team_dv_semi.add(f'C{9+g}')   # Serie A local
    team_dv_semi.add(f'D{9+g}')   # Serie A visitante
    team_dv_semi.add(f'C{21+g}')  # Serie B local
    team_dv_semi.add(f'D{21+g}')  # Serie B visitante

# Team dropdowns for Finals
team_dv_final = DataValidation(type="list", formula1=team_range, allow_blank=True)
ws_final.add_data_validation(team_dv_final)
team_dv_final.add('B3')  # team 1
team_dv_final.add('E3')  # team 2
for g in range(5):
    team_dv_final.add(f'C{7+g}')  # local
    team_dv_final.add(f'D{7+g}')  # visitante

# Save
output = "P:/Games/MVP Cuba 2011/Torneo/Torneo MVP Cuba 2011.xlsx"
wb.save(output)
print(f"Created: {output}")
